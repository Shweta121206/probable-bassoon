"""Excel report generation for bank statement transactions."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.metadata import StatementMetadata

logger = logging.getLogger(__name__)

TRANSACTION_COLUMNS = (
    "Date",
    "Transaction Type",
    "Description",
    "Debit",
    "Credit",
    "Balance",
    "Classification Code",
    "Classification Name",
    "Classification Type",
    "Confidence",
)

SUMMARY_COLUMNS = (
    "Classification Name",
    "Number of Transactions",
    "Total Debit",
    "Total Credit",
    "Net Amount",
)
METADATA_COLUMNS = (
    "Bank Name",
    "Account Holder",
    "Account Number",
    "Statement Start",
    "Statement End",
    "Statement Year",
)
BALANCE_COLUMNS = ("Date", "Balance")


def create_transactions_workbook(
    transactions: list[dict[str, Any]],
    output_path: str,
    metadata: StatementMetadata | None = None,
) -> str:
    """Create a basic Excel workbook containing transaction rows."""
    logger.info("Creating transactions workbook with %d row(s)", len(transactions))

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Transactions"

    _write_transactions_sheet(worksheet, transactions, metadata)
    _format_transactions_sheet(worksheet)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)

    logger.info("Saved transactions workbook: %s", path)
    return str(path)


def create_bank_statement_workbook(
    transactions: list[dict[str, Any]],
    metadata: StatementMetadata | None,
    balance_history: list[dict[str, Any]],
    output_path: str,
) -> str:
    """Create a multi-sheet Excel workbook for the bank statement output."""
    logger.info(
        "Creating bank statement workbook with %d transaction(s), %d balance entry(ies)",
        len(transactions),
        len(balance_history),
    )

    workbook = Workbook()

    transactions_sheet = workbook.active
    transactions_sheet.title = "Transactions"
    _write_transactions_sheet(transactions_sheet, transactions, metadata)
    _format_transactions_sheet(transactions_sheet)

    summary_sheet = workbook.create_sheet("Classification Summary")
    _write_classification_summary_sheet(summary_sheet, transactions)
    _format_summary_sheet(summary_sheet)

    metadata_sheet = workbook.create_sheet("Statement Metadata")
    _write_metadata_sheet(metadata_sheet, metadata)
    _format_metadata_sheet(metadata_sheet)

    balance_sheet = workbook.create_sheet("Balance History")
    _write_balance_sheet(balance_sheet, balance_history)
    _format_balance_sheet(balance_sheet)

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    temp_path = path.with_suffix(".tmp.xlsx")
    workbook.save(temp_path)

    if path.exists():
        path.unlink()

    temp_path.replace(path)

    logger.info("Saved bank statement workbook: %s", path)
    return str(path)


def _write_transactions_sheet(
    worksheet: Worksheet,
    transactions: list[dict[str, Any]],
    metadata: StatementMetadata | None = None,
) -> None:
    """Write transaction headers and rows to a worksheet."""
    worksheet.append(TRANSACTION_COLUMNS)

    for transaction in transactions:
        amount = transaction.get("amount")
        debit = transaction.get("debit")
        credit = transaction.get("credit")
        if debit is None and credit is None and isinstance(amount, (int, float)):
            if amount < 0:
                debit = abs(amount)
                credit = 0
            else:
                debit = 0
                credit = amount

        worksheet.append(
            (
                transaction.get("date", ""),
                transaction.get("transaction_type", ""),
                transaction.get("description", ""),
                debit if debit is not None else 0,
                credit if credit is not None else 0,
                transaction.get("balance", ""),
                transaction.get("classification_code", ""),
                transaction.get("classification_name", ""),
                transaction.get("classification_type", ""),
                transaction.get("classification_confidence", ""),
            )
        )


def _write_metadata_sheet(
    worksheet: Worksheet,
    metadata: StatementMetadata | None,
) -> None:
    """Write statement metadata headers and values to a worksheet."""
    worksheet.append(METADATA_COLUMNS)

    if metadata is None:
        worksheet.append(("", "", "", "", "", ""))
        return

    worksheet.append(
        (
            metadata.bank_name or "",
            metadata.account_holder or "",
            metadata.account_number or "",
            metadata.statement_period_start or "",
            metadata.statement_period_end or "",
            metadata.statement_year if metadata.statement_year is not None else "",
        )
    )


def _write_balance_sheet(
    worksheet: Worksheet,
    balance_history: list[dict[str, Any]],
) -> None:
    """Write balance history headers and rows to a worksheet."""
    worksheet.append(BALANCE_COLUMNS)

    for entry in balance_history:
        worksheet.append((entry.get("date", ""), entry.get("balance", "")))


def _format_transactions_sheet(worksheet: Worksheet) -> None:
    """Apply formatting and autosize columns for the transactions sheet."""
    _apply_header_style(worksheet)
    _apply_table_style(worksheet)

    for cell in worksheet["D"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'
        cell.alignment = Alignment(horizontal="right")

    for cell in worksheet["E"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'
        cell.alignment = Alignment(horizontal="right")

    for cell in worksheet["F"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'
        cell.alignment = Alignment(horizontal="right")

    _autosize_columns(worksheet)


def _format_metadata_sheet(worksheet: Worksheet) -> None:
    """Apply formatting and autosize columns for the metadata sheet."""
    _apply_header_style(worksheet)
    _apply_table_style(worksheet)
    _autosize_columns(worksheet)


def _write_classification_summary_sheet(
    worksheet: Worksheet,
    transactions: list[dict[str, Any]],
) -> None:
    """Write the classification summary sheet grouping by classification name."""
    worksheet.append(SUMMARY_COLUMNS)

    summary: dict[str, dict[str, float | int]] = {}
    for transaction in transactions:
        classification_name = transaction.get("classification_name", "Uncategorized") or "Uncategorized"
        debit = transaction.get("debit")
        credit = transaction.get("credit")
        debit_value = float(debit) if isinstance(debit, (int, float)) else 0.0
        credit_value = float(credit) if isinstance(credit, (int, float)) else 0.0
        entry = summary.setdefault(
            classification_name,
            {
                "count": 0,
                "total_debit": 0.0,
                "total_credit": 0.0,
                "net_amount": 0.0,
            },
        )
        entry["count"] += 1
        entry["total_debit"] += debit_value
        entry["total_credit"] += credit_value
        entry["net_amount"] += credit_value - debit_value

    for classification_name, values in sorted(summary.items()):
        worksheet.append(
            (
                classification_name,
                values["count"],
                values["total_debit"],
                values["total_credit"],
                values["net_amount"],
            )
        )


def _format_summary_sheet(worksheet: Worksheet) -> None:
    _apply_header_style(worksheet)
    _apply_table_style(worksheet)

    for cell in worksheet["C"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'
        cell.alignment = Alignment(horizontal="right")

    for cell in worksheet["D"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'
        cell.alignment = Alignment(horizontal="right")

    for cell in worksheet["E"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'
        cell.alignment = Alignment(horizontal="right")

    _autosize_columns(worksheet)


def _format_balance_sheet(worksheet: Worksheet) -> None:
    """Apply formatting and autosize columns for the balance history sheet."""
    _apply_header_style(worksheet)
    _apply_table_style(worksheet)

    for cell in worksheet["B"][1:]:
        cell.number_format = '#,##0.00;[Red]-#,##0.00'
        cell.alignment = Alignment(horizontal="right")

    _autosize_columns(worksheet)


def _apply_header_style(worksheet: Worksheet) -> None:
    """Style the header row for a worksheet."""
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")


def _apply_table_style(worksheet: Worksheet) -> None:
    """Apply row and alignment styling to a worksheet."""
    thin_border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top")


def _autosize_columns(worksheet: Worksheet) -> None:
    """Autosize worksheet columns based on visible cell values."""
    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)
