"""Chart-of-account retrieval using sentence embeddings."""

from __future__ import annotations

import hashlib
import logging
from pathlib import Path
from typing import Any

import numpy as np
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

DEFAULT_MODEL_NAME = "all-MiniLM-L6-v2"
DEFAULT_ACCOUNT_PATH = Path(__file__).resolve().parents[1] / ".." / "ChartOfAccounts - TEMPLATE.xlsx"


class ChartOfAccountsRetriever:
    """Retrieve the closest chart-of-account entries for a transaction description."""

    def __init__(self, account_path: str | Path | None = None, model_name: str = DEFAULT_MODEL_NAME) -> None:
        self.account_path = Path(account_path) if account_path is not None else DEFAULT_ACCOUNT_PATH
        self.model_name = model_name
        self._accounts: list[dict[str, str]] = []
        self._account_names: list[str] = []
        self._embeddings: np.ndarray | None = None
        self._embedding_cache: dict[str, np.ndarray] = {}
        self._model = None
        self._load_accounts()
        self._initialize_model()
        self._build_account_embeddings()

    def retrieve(self, transaction_description: str, top_n: int = 5) -> list[dict[str, Any]]:
        """Return the top matching accounts for a single transaction description."""
        if top_n <= 0:
            return []

        if self._model is None:
            return self._retrieve_with_fallback(transaction_description, top_n)

        embedding = self._get_embedding(transaction_description)
        if self._embeddings is None or self._embeddings.size == 0:
            return []

        similarities = self._cosine_similarity(embedding, self._embeddings)
        similarity_scores = similarities[0] if similarities.ndim > 1 else similarities
        ranked_indices = np.argsort(similarity_scores)[::-1][: min(top_n, len(self._accounts))]

        results: list[dict[str, Any]] = []
        for index in ranked_indices:
            account = self._accounts[int(index)]
            results.append(
                {
                    "account_code": account.get("account_code", ""),
                    "account_name": account.get("account_name", ""),
                    "account_type": account.get("account_type", ""),
                    "similarity": float(similarity_scores[int(index)]),
                }
            )
        return results

    def retrieve_batch(self, transaction_descriptions: list[str], top_n: int = 5) -> list[list[dict[str, Any]]]:
        """Return ranked matches for each description in a batch."""
        return [self.retrieve(description, top_n=top_n) for description in transaction_descriptions]

    def _load_accounts(self) -> None:
        if not self.account_path.exists():
            logger.warning("Chart of accounts workbook not found at %s", self.account_path)
            self._accounts = []
            self._account_names = []
            return

        workbook = load_workbook(self.account_path, read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                self._accounts = []
                self._account_names = []
                return

            headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
            code_index = headers.index("*code") if "*code" in headers else -1
            name_index = headers.index("*name") if "*name" in headers else -1
            type_index = headers.index("*type") if "*type" in headers else -1

            self._accounts = []
            for row in rows[1:]:
                if not row:
                    continue

                account_code = row[code_index] if code_index >= 0 and code_index < len(row) else ""
                account_name = row[name_index] if name_index >= 0 and name_index < len(row) else ""
                account_type = row[type_index] if type_index >= 0 and type_index < len(row) else ""

                if not account_name:
                    continue

                self._accounts.append(
                    {
                        "account_code": str(account_code) if account_code is not None else "",
                        "account_name": str(account_name),
                        "account_type": str(account_type) if account_type is not None else "",
                    }
                )

            self._account_names = [account.get("account_name", "") for account in self._accounts]
        finally:
            workbook.close()

    def _initialize_model(self) -> None:
        try:
            from sentence_transformers import SentenceTransformer
        except ImportError:
            logger.warning("sentence-transformers is not installed; using a lightweight fallback encoder")
            self._model = None
            return

        self._model = SentenceTransformer(self.model_name)

    def _build_account_embeddings(self) -> None:
        if not self._account_names:
            self._embeddings = np.array([], dtype=float)
            return

        if self._model is None:
            self._embeddings = np.array([self._simple_embedding(name) for name in self._account_names], dtype=float)
            return

        try:
            self._embeddings = np.array(self._model.encode(self._account_names, convert_to_numpy=True), dtype=float)
        except Exception as exc:  # pragma: no cover - defensive fallback
            logger.warning("Embedding generation failed with sentence-transformers: %s", exc)
            self._embeddings = np.array([self._simple_embedding(name) for name in self._account_names], dtype=float)

    def _get_embedding(self, text: str) -> np.ndarray:
        cache_key = self._cache_key(text)
        if cache_key in self._embedding_cache:
            return self._embedding_cache[cache_key]

        if self._model is None:
            embedding = self._simple_embedding(text)
        else:
            try:
                embedding = self._model.encode([text], convert_to_numpy=True)[0]
            except Exception as exc:  # pragma: no cover - defensive fallback
                logger.warning("Embedding generation failed for text %s: %s", text, exc)
                embedding = self._simple_embedding(text)

        embedding_array = np.array(embedding, dtype=float)
        self._embedding_cache[cache_key] = embedding_array
        return embedding_array

    def _retrieve_with_fallback(self, transaction_description: str, top_n: int) -> list[dict[str, Any]]:
        description_tokens = self._tokenize(transaction_description)
        scored_accounts: list[tuple[float, dict[str, Any]]] = []

        for account in self._accounts:
            account_name = account.get("account_name", "")
            account_tokens = self._tokenize(account_name)
            score = self._token_overlap_score(description_tokens, account_tokens)
            scored_accounts.append((score, account))

        scored_accounts.sort(key=lambda item: (-item[0], item[1].get("account_name", "")))
        ranked_accounts = scored_accounts[: min(top_n, len(scored_accounts))]

        return [
            {
                "account_code": account.get("account_code", ""),
                "account_name": account.get("account_name", ""),
                "account_type": account.get("account_type", ""),
                "similarity": round(score, 3),
            }
            for score, account in ranked_accounts
        ]

    @staticmethod
    def _tokenize(text: str) -> set[str]:
        tokens = {
            ChartOfAccountsRetriever._normalize_token(token)
            for token in text.lower().replace("/", " ").split()
            if token
        }
        return {token for token in tokens if token}

    @staticmethod
    def _normalize_token(token: str) -> str:
        cleaned = "".join(ch for ch in token if ch.isalnum())
        if len(cleaned) <= 2:
            return cleaned
        if cleaned.endswith("ies") and len(cleaned) > 4:
            return cleaned[:-3] + "y"
        if cleaned.endswith("s") and len(cleaned) > 3:
            return cleaned[:-1]
        return cleaned

    @staticmethod
    def _token_overlap_score(description_tokens: set[str], account_tokens: set[str]) -> float:
        if not description_tokens or not account_tokens:
            return 0.0

        overlap = description_tokens & account_tokens
        if overlap:
            return 1.0 + 0.2 * len(overlap)

        description_terms = sorted(description_tokens)
        account_terms = sorted(account_tokens)
        shared_prefixes = 0
        for description_term in description_terms:
            for account_term in account_terms:
                if description_term.startswith(account_term) or account_term.startswith(description_term):
                    shared_prefixes += 0.1
                    break
        return shared_prefixes

    def _cosine_similarity(self, vector_a: np.ndarray, vector_b: np.ndarray) -> np.ndarray:
        vector_a = vector_a.astype(float)
        vector_b = vector_b.astype(float)
        if vector_a.ndim == 1:
            vector_a = vector_a.reshape(1, -1)
        if vector_b.ndim == 1:
            vector_b = vector_b.reshape(1, -1)

        vector_a_norm = np.linalg.norm(vector_a, axis=1, keepdims=True)
        vector_b_norm = np.linalg.norm(vector_b, axis=1, keepdims=True)
        numerator = vector_a @ vector_b.T
        denominator = np.maximum(vector_a_norm * vector_b_norm.T, 1e-12)
        return numerator / denominator

    @staticmethod
    def _cache_key(text: str) -> str:
        return hashlib.sha256(text.lower().strip().encode("utf-8")).hexdigest()

    @staticmethod
    def _simple_embedding(text: str) -> np.ndarray:
        tokens = [token for token in text.lower().split() if token]
        if not tokens:
            return np.zeros(32, dtype=float)

        vector = np.zeros(32, dtype=float)
        for token_index, token in enumerate(tokens):
            digest = hashlib.md5(token.encode("utf-8")).hexdigest()
            value = int(digest[:8], 16) / 0xFFFFFFFF
            vector += np.linspace(0.1, 1.0, 32) * value * (1.0 / (token_index + 1))

        return vector / max(len(tokens), 1)
